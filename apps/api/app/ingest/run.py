from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.config import Settings
from app.db.models import Account, Document, Order, Ticket
from app.ingest.document_registry import get_document_registry
from app.timeutil import parse_assessment_datetime


def _sheet_rows(path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: list[dict] = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {
            headers[i]: row[i] for i in range(len(headers)) if i < len(row)
        }
        records.append(record)
    return records


def _as_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes"}


def ingest_excel(session: Session, settings: Settings) -> dict[str, int]:
    excel_path = settings.resolved_data_path / "ParcelPilot_Assessment_Data.xlsx"
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel data pack not found: {excel_path}")

    counts = {"accounts": 0, "orders": 0, "tickets": 0}

    for row in _sheet_rows(excel_path, "accounts"):
        account = Account(
            account_id=str(row["account_id"]),
            account_name=str(row["account_name"]),
            plan=str(row["plan"]),
            status=str(row["status"]),
            csm=row.get("csm"),
            contract_file=row.get("contract_file"),
            premium_support=_as_bool(row.get("premium_support")),
            notes=row.get("notes"),
        )
        session.merge(account)
        counts["accounts"] += 1

    for row in _sheet_rows(excel_path, "orders"):
        order = Order(
            order_id=str(row["order_id"]),
            account_id=str(row["account_id"]),
            carrier=str(row["carrier"]),
            status=str(row["status"]),
            booked_at=parse_assessment_datetime(row["booked_at"]),
            pickup_window_start=parse_assessment_datetime(row["pickup_window_start"]),
            pickup_window_end=parse_assessment_datetime(row["pickup_window_end"]),
            pickup_actual_at=parse_assessment_datetime(row.get("pickup_actual_at")),
            shipment_fee_inr=float(row["shipment_fee_inr"]),
            carrier_fault=_as_bool(row.get("carrier_fault")),
            customer_fault=_as_bool(row.get("customer_fault")),
            cancellation_requested_at=parse_assessment_datetime(
                row.get("cancellation_requested_at")
            ),
            notes=row.get("notes"),
        )
        session.merge(order)
        counts["orders"] += 1

    for row in _sheet_rows(excel_path, "tickets"):
        ticket = Ticket(
            ticket_id=str(row["ticket_id"]),
            account_id=str(row["account_id"]),
            created_at=parse_assessment_datetime(row["created_at"]),
            status=str(row["status"]),
            subject=str(row["subject"]),
            description=row.get("description"),
            channel=row.get("channel"),
            assigned_to=row.get("assigned_to"),
            last_customer_message_at=parse_assessment_datetime(
                row.get("last_customer_message_at")
            ),
            historical_resolution=row.get("historical_resolution"),
        )
        session.merge(ticket)
        counts["tickets"] += 1

    return counts


def ingest_documents(session: Session) -> int:
    count = 0
    for entry in get_document_registry():
        document = Document(**entry)
        session.merge(document)
        count += 1
    return count


def run_ingest(settings: Settings) -> dict[str, int | str]:
    from app.db.schema import create_all_tables
    from app.db.session import get_session_factory

    create_all_tables()
    session = get_session_factory()()
    try:
        excel_counts = ingest_excel(session, settings)
        doc_count = ingest_documents(session)
        session.commit()
        return {
            **excel_counts,
            "documents": doc_count,
            "snapshot_at": settings.snapshot_at,
        }
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
